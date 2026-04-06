import pandas as pd
import io
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
# Rows 1 and 2 are empty string or nothing
ws.cell(row=3, column=1, value="Head1")
ws.cell(row=3, column=2, value="Head2")
ws.cell(row=4, column=1, value="Data1")
ws.cell(row=4, column=2, value="Data2")

bio = io.BytesIO()
wb.save(bio)
bio.seek(0)

df = pd.read_excel(bio, header=0)
print("Columns:", df.columns.tolist())
print(df)
